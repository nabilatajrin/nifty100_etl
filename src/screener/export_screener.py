"""Generate output/screener_output.xlsx (Sprint 3, Day 17).

One sheet per preset, sorted by composite score descending, with cells
colour-coded green (meets threshold) / red (fails threshold).
Run:  python -m src.screener.export_screener
"""

import os
import sqlite3
import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from .engine import load_config, run_preset
from .composite_score import composite_score
from .run_presets import latest_ratios, PRESETS

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Arial")
BODY_FONT = Font(name="Arial")

# columns to show in each sheet
DISPLAY_COLS = [
    "company_id", "composite_quality_score", "return_on_equity_pct",
    "operating_profit_margin_pct", "debt_to_equity", "interest_coverage",
    "free_cash_flow_cr", "revenue_cagr_5yr", "pat_cagr_5yr",
    "pe_ratio", "pb_ratio", "dividend_yield_pct",
]


def _passes(metric_key, value, threshold, config):
    """Does a single cell value meet the preset threshold?"""
    if pd.isna(value):
        return None
    meta = config["metrics"].get(metric_key)
    if not meta:
        return None
    if meta["direction"] == "min":
        return value >= threshold
    return value <= threshold


# map display column -> metric key in config
COL_TO_METRIC = {
    "return_on_equity_pct": "roe",
    "operating_profit_margin_pct": "opm",
    "debt_to_equity": "de",
    "interest_coverage": "icr",
    "free_cash_flow_cr": "fcf",
    "revenue_cagr_5yr": "rev_cagr_5yr",
    "pat_cagr_5yr": "pat_cagr_5yr",
    "pe_ratio": "pe",
    "pb_ratio": "pb",
    "dividend_yield_pct": "dividend_yield",
}


def write_sheet(ws, df, preset_thresholds, config):
    cols = [c for c in DISPLAY_COLS if c in df.columns]

    # header
    for j, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.fill = HEADER
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # body
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        for j, col in enumerate(cols, 1):
            val = row[col]
            cell = ws.cell(row=i, column=j,
                           value=round(val, 2) if isinstance(val, (int, float)) and pd.notna(val) else val)
            cell.font = BODY_FONT
            metric_key = COL_TO_METRIC.get(col)
            if metric_key and metric_key in preset_thresholds:
                res = _passes(metric_key, val, preset_thresholds[metric_key], config)
                if res is True:
                    cell.fill = GREEN
                elif res is False:
                    cell.fill = RED

    # column widths
    for j, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = max(12, len(col) + 2)


def main():
    load_dotenv()
    db = os.getenv("DB_PATH", "data/nifty100.db")
    conn = sqlite3.connect(db)
    df = latest_ratios(conn)
    sectors_df = pd.read_sql("SELECT company_id, broad_sector FROM sectors", conn)
    conn.close()

    # attach sector + composite score
    df = df.merge(sectors_df, on="company_id", how="left")
    df["composite_quality_score"] = composite_score(df)

    sectors = df["broad_sector"]
    sectors.index = df.index
    config = load_config()

    wb = Workbook()
    wb.remove(wb.active)

    for preset in PRESETS:
        result = run_preset(df, preset, config, sectors)
        result = result.sort_values("composite_quality_score", ascending=False)
        ws = wb.create_sheet(title=preset[:31])  # Excel 31-char sheet limit
        write_sheet(ws, result, config["presets"][preset], config)
        print(f"{preset}: {len(result)} rows")

    out_dir = "output"
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "screener_output.xlsx")
    wb.save(out_path)
    print(f"\nSaved {out_path}")


if __name__ == "__main__":
    main()
