"""Generate output/peer_comparison.xlsx (Sprint 3, Day 20).

11 sheets, one per peer group. Each sheet: company_id, company_name, metric
values, and percentile-rank columns colour-coded green (>=75th), yellow
(25th-75th), red (<=25th). Benchmark company row highlighted gold/amber.
Summary row at the bottom shows the peer-group median for each metric.

Run:  python -m src.reports.export_peer_comparison
"""

import os
import sqlite3
import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOLD = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
HEADER = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Arial")
BODY_FONT = Font(name="Arial")
BOLD_FONT = Font(name="Arial", bold=True)

METRICS = [
    "return_on_equity_pct", "operating_profit_margin_pct", "net_profit_margin_pct",
    "debt_to_equity", "free_cash_flow_cr", "pat_cagr_5yr", "revenue_cagr_5yr",
    "eps_cagr_5yr", "interest_coverage", "asset_turnover",
]


def _pct_fill(pct_rank: float):
    if pct_rank is None:
        return None
    if pct_rank >= 0.75:
        return GREEN
    if pct_rank <= 0.25:
        return RED
    return YELLOW


def load_data(conn):
    pct = pd.read_sql("SELECT * FROM peer_percentiles", conn)
    companies = pd.read_sql("SELECT id AS company_id, company_name FROM companies", conn)

    peers = pd.read_sql("SELECT * FROM peer_groups", conn)
    group_col = next((c for c in peers.columns
                      if "group" in c.lower() and "name" in c.lower()), None)
    if group_col is None:
        group_col = next((c for c in peers.columns
                          if "peer" in c.lower() and c != "company_id"), None)
    bench_col = next((c for c in peers.columns if "benchmark" in c.lower()), None)
    peers = peers.rename(columns={group_col: "peer_group_name"})
    return pct, companies, peers, bench_col


def build_wide(pct: pd.DataFrame, group: str) -> pd.DataFrame:
    """Pivot the long percentile table to wide: one row per company, cols = metrics."""
    sub = pct[pct["peer_group_name"] == group]
    values = sub.pivot_table(index="company_id", columns="metric", values="value", aggfunc="first")
    ranks = sub.pivot_table(index="company_id", columns="metric", values="percentile_rank", aggfunc="first")
    values.columns = [f"{c}" for c in values.columns]
    ranks.columns = [f"{c}_pctile" for c in ranks.columns]
    return values.join(ranks).reset_index()


def write_sheet(ws, wide: pd.DataFrame, companies: pd.DataFrame,
                benchmark_company: str | None):
    wide = wide.merge(companies, on="company_id", how="left")

    metric_cols = [c for c in METRICS if c in wide.columns]
    cols = ["company_id", "company_name"] + [c for pair in metric_cols
            for c in (pair, f"{pair}_pctile") if c in wide.columns]

    # header
    for j, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.fill = HEADER
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # body
    for i, (_, row) in enumerate(wide.iterrows(), start=2):
        is_bench = benchmark_company and row["company_id"] == benchmark_company
        for j, col in enumerate(cols, 1):
            val = row.get(col)
            display = round(val, 2) if isinstance(val, (int, float)) and pd.notna(val) else val
            cell = ws.cell(row=i, column=j, value=display)
            cell.font = BOLD_FONT if is_bench else BODY_FONT
            if is_bench:
                cell.fill = GOLD
            elif col.endswith("_pctile") and isinstance(val, (int, float)) and pd.notna(val):
                fill = _pct_fill(val)
                if fill:
                    cell.fill = fill

    # summary row: peer group median for each metric
    summary_row = len(wide) + 2
    ws.cell(row=summary_row, column=1, value="Peer Group Median").font = BOLD_FONT
    for j, col in enumerate(cols, 1):
        if col in wide.columns and col not in ("company_id", "company_name"):
            median_val = pd.to_numeric(wide[col], errors="coerce").median()
            if pd.notna(median_val):
                ws.cell(row=summary_row, column=j, value=round(median_val, 2)).font = BOLD_FONT

    for j, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = max(14, len(col) + 2)


def main():
    load_dotenv()
    db = os.getenv("DB_PATH", "data/nifty100.db")
    conn = sqlite3.connect(db)
    pct, companies, peers, bench_col = load_data(conn)
    conn.close()

    groups = sorted(pct["peer_group_name"].dropna().unique())
    print(f"Peer groups found: {len(groups)}")

    wb = Workbook()
    wb.remove(wb.active)

    for group in groups:
        wide = build_wide(pct, group)

        benchmark_company = None
        if bench_col:
            grp_rows = peers[peers["peer_group_name"] == group]
            true_rows = grp_rows[grp_rows[bench_col].astype(str).str.upper().isin(["TRUE", "1", "YES"])]
            if not true_rows.empty:
                benchmark_company = true_rows.iloc[0]["company_id"]

        ws = wb.create_sheet(title=str(group)[:31])
        write_sheet(ws, wide, companies, benchmark_company)
        print(f"  {group}: {len(wide)} companies" +
              (f" (benchmark: {benchmark_company})" if benchmark_company else ""))

    out_dir = "output"
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "peer_comparison.xlsx")
    wb.save(out_path)
    print(f"\nSaved {out_path} with {len(wb.sheetnames)} sheets")


if __name__ == "__main__":
    main()
